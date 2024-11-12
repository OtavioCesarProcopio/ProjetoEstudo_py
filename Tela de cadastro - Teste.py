import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearance()
        self.todo_sistema()

    def layout_config(self):
        self.title("Tela de cadastro")
        self.geometry("700x500")

    def appearance(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', '#fff']).place(x=50, y=440)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_apm).place(x=50, y=470)

    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=10, bg_color="teal", fg_color="teal")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Cadastro de Cliente", font=("Century Gothic bold", 24), text_color="#fff").place(x= 190, y=15)

        span = ctk.CTkLabel(self, text="Por favor, preecha os campos", font=("Century Gothic bold", 16), text_color=["#000","#fff"]).place(x=50, y=70)

        def submit():

            ficheiro = pathlib.Path("Clientes.xlsx")

            if ficheiro.exists():
                pass
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Cadastros"
                sheet.append(["Nome", "Contato", "Idade", "Endereço","Genero","Observacoes"])  # Cabeçalhos
                workbook.save(ficheiro)

            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            address = address_value.get()
            gender = gender_combobox.get()
            obs = obs_entry.get(0.0, END)

            if(name =="" or contact =="" or age =="" or address =="" or gender ==""):
                messagebox.showerror("Erro", "Por favor, preencha todos os campos")
            else:
                ficheiro = openpyxl.load_workbook('Clientes.xlsx')
                folha = ficheiro.active
                folha.cell(column=1, row=folha.max_row+1, value=name)
                folha.cell(column=2, row=folha.max_row, value=contact)
                folha.cell(column=3, row=folha.max_row, value=age)
                folha.cell(column=4, row=folha.max_row, value=address)
                folha.cell(column=5, row=folha.max_row, value=gender)
                folha.cell(column=6, row=folha.max_row, value=obs)

                ficheiro.save(r"Clientes.xlsx")
                messagebox.showinfo("Sistema", "Dados salvos com sucesso!")
                clear()

        def clear():
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            address_value.set("")
            obs_entry.delete(0.0, END)

        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()
        

        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic bold", 16), fg_color="transparent")
        contact_entry = ctk.CTkEntry(self, width=200, textvariable=contact_value, font=("Century Gothic bold", 16), fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width=150, textvariable=age_value, font=("Century Gothic bold", 16), fg_color="transparent")
        address_entry = ctk.CTkEntry(self, width=200, textvariable=address_value, font=("Century Gothic bold", 16), fg_color="transparent")

        gender_combobox = ctk.CTkComboBox(self, values=("Masculino","Feminino"), font=("Century Gothic bold", 14), width=150)        
        gender_combobox.set("Masculino")

        obs_entry = ctk.CTkTextbox(self, width=600, height=150, font=("arial", 19), border_color="#aaa", border_width=2, fg_color="transparent")

        lb_name = ctk.CTkLabel(self, text="Nome", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_contact = ctk.CTkLabel(self, text="Contato", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_age = ctk.CTkLabel(self, text="Idade", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_gender = ctk.CTkLabel(self, text="Genero", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_address = ctk.CTkLabel(self, text="Endereço", font=("Century Gothic bold", 16), text_color=["#000","#fff"])
        lb_obs = ctk.CTkLabel(self, text="Observação", font=("Century Gothic bold", 16), text_color=["#000","#fff"])

        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131")
        btn_clear = ctk.CTkButton(self, text="Limpar campos".upper(), command=clear, fg_color="#555", hover_color="#333")

        btn_submit.place(x=300, y=470)
        btn_clear.place(x=510, y=470)


        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contact.place(x=450, y=120)
        contact_entry.place(x=450, y=150)

        lb_address.place(x=50 , y=190)
        address_entry.place(x=50 , y=220)

        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)

        lb_gender.place(x=500 , y=190)
        gender_combobox.place(x=500 , y=220)

        lb_obs.place(x=50 , y=260)
        obs_entry.place(x=50 , y=290)

          

if __name__ == "__main__":
    app = App()
    app.mainloop()